"""
Revenue & Transactions API Router.

Provides endpoints for:
- Listing transactions with filters & pagination
- Revenue summary aggregation
- Revenue grouped by state
- Monthly revenue (last 6 months)
- Payment-method distribution
- Recharge-amount distribution
- Exporting revenue reports (Excel / PDF)
"""

import io
import calendar
from datetime import datetime, timedelta
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, case
from sqlalchemy.orm import Session

from config.database import get_db
from models.transaction import Transaction
from models.user import User
from schemas.revenue_schema import (
    PaginatedTransactions,
    RevenueSummary,
    RevenueByState,
    MonthlyRevenue,
    PaymentMethodDistribution,
    RechargeDistribution,
    TransactionResponse,
)

router = APIRouter(prefix="/api", tags=["Revenue & Transactions"])


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/transactions
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/transactions", response_model=PaginatedTransactions)
def list_transactions(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Items per page"),
    date_from: Optional[datetime] = Query(None, description="Filter from date (ISO 8601)"),
    date_to: Optional[datetime] = Query(None, description="Filter to date (ISO 8601)"),
    payment_method: Optional[str] = Query(None, description="Filter by payment method"),
    status: Optional[str] = Query(None, description="Filter by status"),
    db: Session = Depends(get_db),
):
    """
    Return a paginated list of all recharge / payment transactions.

    Supports optional filters:
    - date_from / date_to  — date range filter
    - payment_method       — e.g. UPI, Net Banking, Credit Card
    - status               — Success | Pending | Failed
    """
    query = db.query(Transaction)

    if date_from:
        query = query.filter(Transaction.created_at >= date_from)
    if date_to:
        query = query.filter(Transaction.created_at <= date_to)
    if payment_method:
        query = query.filter(Transaction.payment_method == payment_method)
    if status:
        query = query.filter(Transaction.status == status)

    total = query.count()
    records = (
        query.order_by(Transaction.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return PaginatedTransactions(
        total=total,
        page=page,
        page_size=page_size,
        data=[TransactionResponse.model_validate(r) for r in records],
    )


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/revenue/summary
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/revenue/summary", response_model=RevenueSummary)
def revenue_summary(db: Session = Depends(get_db)):
    """
    Return aggregated revenue statistics:

    - total_revenue      — sum of all successful transactions
    - pending_revenue    — sum of pending transactions
    - average_bill_amount — average bill_amount across all users
    - highest_payment    — highest single transaction amount
    """
    total_revenue = (
        db.query(func.coalesce(func.sum(Transaction.amount), 0))
        .filter(Transaction.status == "Success")
        .scalar()
    )
    pending_revenue = (
        db.query(func.coalesce(func.sum(Transaction.amount), 0))
        .filter(Transaction.status == "Pending")
        .scalar()
    )
    avg_bill = db.query(func.coalesce(func.avg(User.bill_amount), 0)).scalar()
    highest_payment = (
        db.query(func.coalesce(func.max(Transaction.amount), 0)).scalar()
    )

    return RevenueSummary(
        total_revenue=float(total_revenue),
        pending_revenue=float(pending_revenue),
        average_bill_amount=float(avg_bill),
        highest_payment=float(highest_payment),
    )


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/revenue/by-state
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/revenue/by-state", response_model=List[RevenueByState])
def revenue_by_state(db: Session = Depends(get_db)):
    """
    Return total successful revenue grouped by state.

    Revenue is pulled from the transactions table; the state column
    is denormalised onto each transaction row for fast aggregation.
    """
    rows = (
        db.query(
            Transaction.state,
            func.coalesce(func.sum(Transaction.amount), 0).label("revenue"),
        )
        .filter(Transaction.status == "Success", Transaction.state.isnot(None))
        .group_by(Transaction.state)
        .order_by(func.sum(Transaction.amount).desc())
        .all()
    )

    return [RevenueByState(state=row.state, revenue=float(row.revenue)) for row in rows]


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/revenue/monthly
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/revenue/monthly", response_model=List[MonthlyRevenue])
def monthly_revenue(db: Session = Depends(get_db)):
    """
    Return monthly revenue data for the last 6 months.

    Only successful transactions are counted.
    """
    now = datetime.utcnow()
    results = []

    for offset in range(5, -1, -1):
        # Calculate the first and last day of the target month
        target = now - timedelta(days=offset * 30)
        year, month = target.year, target.month
        first_day = datetime(year, month, 1)
        last_day = datetime(year, month, calendar.monthrange(year, month)[1], 23, 59, 59)

        revenue = (
            db.query(func.coalesce(func.sum(Transaction.amount), 0))
            .filter(
                Transaction.status == "Success",
                Transaction.created_at >= first_day,
                Transaction.created_at <= last_day,
            )
            .scalar()
        )

        results.append(
            MonthlyRevenue(
                month=first_day.strftime("%b"),
                revenue=float(revenue),
            )
        )

    return results


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/revenue/payment-methods
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/revenue/payment-methods", response_model=List[PaymentMethodDistribution])
def payment_method_distribution(db: Session = Depends(get_db)):
    """
    Return count and total amount grouped by payment method.

    Covers: UPI, Net Banking, Credit Card, Debit Card, Wallet.
    """
    rows = (
        db.query(
            Transaction.payment_method,
            func.count(Transaction.id).label("count"),
            func.coalesce(func.sum(Transaction.amount), 0).label("total_amount"),
        )
        .group_by(Transaction.payment_method)
        .all()
    )

    return [
        PaymentMethodDistribution(
            payment_method=row.payment_method,
            count=row.count,
            total_amount=float(row.total_amount),
        )
        for row in rows
    ]


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/revenue/recharge-distribution
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/revenue/recharge-distribution", response_model=List[RechargeDistribution])
def recharge_distribution(db: Session = Depends(get_db)):
    """
    Return count of transactions bucketed into recharge-amount ranges.

    Buckets: 0-100 | 100-500 | 500-1000 | 1000+
    """
    buckets = {
        "0-100": (0, 100),
        "100-500": (100, 500),
        "500-1000": (500, 1000),
        "1000+": (1000, None),
    }

    results = []
    for label, (low, high) in buckets.items():
        q = db.query(func.count(Transaction.id)).filter(Transaction.amount >= low)
        if high is not None:
            q = q.filter(Transaction.amount < high)
        count = q.scalar() or 0
        results.append(RechargeDistribution(range_label=label, count=count))

    return results


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/revenue/export
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/revenue/export")
def export_revenue(
    format: str = Query("excel", description="Export format: excel | pdf"),
    db: Session = Depends(get_db),
):
    """
    Export revenue report as Excel (.xlsx) or PDF.

    Use query parameter `format=excel` (default) or `format=pdf`.
    """
    transactions = (
        db.query(Transaction)
        .filter(Transaction.status == "Success")
        .order_by(Transaction.created_at.desc())
        .all()
    )

    rows = [
        {
            "Transaction ID": t.transaction_id,
            "User ID": t.user_id,
            "State": t.state or "",
            "Amount (₹)": t.amount,
            "Payment Method": t.payment_method,
            "Status": t.status,
            "Date": t.created_at.strftime("%Y-%m-%d %H:%M"),
        }
        for t in transactions
    ]

    if format.lower() == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="openpyxl is not installed. Run: pip install openpyxl",
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Revenue Report"

        # Header row
        headers = list(rows[0].keys()) if rows else [
            "Transaction ID", "User ID", "State",
            "Amount (₹)", "Payment Method", "Status", "Date",
        ]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = 20

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row.values(), start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=revenue_report.xlsx"},
        )

    elif format.lower() == "pdf":
        try:
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="reportlab is not installed. Run: pip install reportlab",
            )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()

        headers = [
            "Transaction ID", "User ID", "State",
            "Amount (₹)", "Payment Method", "Status", "Date",
        ]
        data = [headers] + [[str(v) for v in row.values()] for row in rows]

        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF3FB")]),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )

        title = Paragraph("WattWise Revenue Report", styles["Title"])
        doc.build([title, table])
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=revenue_report.pdf"},
        )

    else:
        raise HTTPException(
            status_code=400,
            detail="Invalid format. Use 'excel' or 'pdf'.",
        )

